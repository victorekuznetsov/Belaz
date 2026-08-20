# -*- coding: utf-8 -*-
"""Нормализация парка техники БЕЛАЗ из выгрузки SAP «Belaz Fleet with engine.xlsx».

В выгрузке две разновидности строк:
  * «Самосвалы карьерные» — сама машина, серийный номер = VIN;
  * «Узлы. ДВС»          — двигатель, машина указана в «ЕО Вышестоящая ЕО».
Скрипт сводит их в один реестр машин с привязанными двигателями и отдельно
собирает двигатели, у которых машина не указана (резерв либо не-БЕЛАЗ техника).
"""
import json, re, sys
import openpyxl

SRC = "Эксплуатация/Парк/Belaz Fleet with engine.xlsx"
OUT = "Эксплуатация/Парк/fleet.json"

NAME = re.compile(
    r'^(?P<type>.+?)\s+(?P<brand>БЕЛАЗ|БелАЗ|ПЩК)-(?P<model>\d{4,5}[A-ZА-Я]?)\s*№\s*(?P<gar>\S+)\s*$')

EMPTY = {'#', 'Н/Д', 'Не присвоено', 'None', ''}


def cl(v):
    s = '' if v is None else str(v).strip()
    return '' if s in EMPTY else s


def parse_name(s):
    m = NAME.match(s or '')
    if not m:
        return None
    return (m.group('type').strip(), m.group('brand').upper(),
            m.group('model').upper(), m.group('gar'))


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    machines, loose = {}, []
    for r in rows[1:]:
        if not r or not r[1]:
            continue
        owner, parent, pid, park, cls = cl(r[0]), cl(r[1]), cl(r[2]), cl(r[3]), cl(r[5])
        maker, eid, eo, year, serial, mark = cl(r[6]), cl(r[7]), cl(r[8]), cl(r[9]), cl(r[11]), cl(r[13])

        if cls == 'Узлы. ДВС':
            eng = {'model': mark or eo, 'sn': serial, 'maker': maker, 'year': year, 'id': eid}
            p = parse_name(parent)
            if not p:
                loose.append(dict(eng, owner=owner, park=park))
                continue
            typ, brand, model, gar = p
            rec = machines.setdefault((owner, brand, model, gar), {})
            rec.update(owner=owner, brand=brand, model=model, gar=gar, type=typ, park=park, id=pid)
            rec.setdefault('engines', []).append(eng)
        else:
            p = parse_name(eo)
            if not p:
                loose.append({'raw': eo, 'owner': owner, 'park': park})
                continue
            typ, brand, model, gar = p
            rec = machines.setdefault((owner, brand, model, gar), {})
            rec.update(owner=owner, brand=brand, model=model, gar=gar, type=typ, park=park,
                       id=eid, vin=serial, year=year)
            rec.setdefault('engines', [])

    out = {
        'source': SRC,
        'machines': sorted(machines.values(), key=lambda v: (v['owner'], v['model'], v['gar'])),
        'engines_unassigned': loose,
    }
    with open(OUT, 'w', encoding='utf-8') as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)
    print(f"машин: {len(out['machines'])}, ДВС без машины: {len(loose)} -> {OUT}")
    return out


if __name__ == '__main__':
    main()
