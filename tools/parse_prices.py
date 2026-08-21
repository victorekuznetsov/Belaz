# -*- coding: utf-8 -*-
"""Прайс-листы БЕЛАЗ → data/prices.js.

Четыре прайса — четыре площадки с разным назначением, поэтому цены хранятся
отдельными колонками и в каталоге показываются рядом, а не сводятся в одну.
Ключ — Код ОКП; дополнительно строится индекс по Обозначению (заводскому
номеру), чтобы цена находилась и для позиций без ОКП.
"""
import argparse, json, os, re
import openpyxl

# id, метка, файл, строка заголовка, колонки (окп, обозначение, наименование, цена)
SOURCES = [
    ('pk', 'ПК', 'Полюс Красноярск',
     'Прайсы/Прайс лист с ценами 01.09.2025 (стим финал) Красноярск.xlsx',
     6, dict(okp=1, des=2, name=3, price=5)),
    ('pa', 'ПА', 'Полюс Алдан',
     'Прайсы/Перечень зч для прейскуранта кс БЕЛАЗ-7555В итог Алдан.xlsx',
     0, dict(okp=1, des=2, name=3, price=4)),
    ('sl', 'СЛ', 'Сухой Лог',
     'Прайсы/Прайс лист с ценами 01.09.2025 (стрим финал) ГПФК Иркутск.xlsx',
     6, dict(okp=1, des=2, name=3, price=5)),
    ('zak', 'Закупка', 'самостоятельная закупка ЗЧ',
     'Прайсы/Прайс-лист к КП_RA-RZ-2026-001659.xlsx',
     7, dict(okp=1, des=3, name=2, price=4)),
]

RE_OKP = re.compile(r'^\d{10}$')


def cell(v):
    return '' if v is None else str(v).strip()


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def norm_des(s):
    """«7548-1001006», «7548 -1001006» → единый вид без пробелов."""
    return re.sub(r'\s+', '', (s or '')).upper()


def read(path, header_row, cols):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == header_row:
            header = cell(r[cols['price']])
            continue
        if header is None or not r:
            continue
        price = num(r[cols['price']] if cols['price'] < len(r) else None)
        if price is None:
            continue
        okp = cell(r[cols['okp']] if cols['okp'] < len(r) else '')
        des = cell(r[cols['des']] if cols['des'] < len(r) else '')
        name = cell(r[cols['name']] if cols['name'] < len(r) else '')
        rows.append((okp, des, name, price))
    wb.close()
    return header, rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--src', default='.', help='корень с папкой Прайсы/')
    ap.add_argument('--out', default='.', help='корень каталога (куда класть data/)')
    a = ap.parse_args()

    columns, by_okp, by_des = [], {}, {}
    n = len(SOURCES)
    for idx, (cid, label, title, rel, hrow, cols) in enumerate(SOURCES):
        path = os.path.join(a.src, rel)
        header, rows = read(path, hrow, cols)
        columns.append({'id': cid, 'label': label, 'title': title,
                        'file': os.path.basename(rel), 'note': header})
        for okp, des, _name, price in rows:
            for key, store in ((okp, by_okp), (norm_des(des), by_des)):
                if not key:
                    continue
                if store is by_okp and not RE_OKP.match(key):
                    continue
                store.setdefault(key, [None] * n)[idx] = price
        print(f"   {label}: {len(rows)} строк с ценой ({os.path.basename(rel)})")

    out_dir = os.path.join(a.out, 'data')
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, 'prices.js')
    with open(out, 'w', encoding='utf-8') as fh:
        fh.write('/* Цены по площадкам. Порядок значений = порядок PRICES.columns. */\n')
        fh.write('window.PRICES = ')
        json.dump({'currency': 'USD', 'columns': columns, 'okp': by_okp, 'des': by_des},
                  fh, ensure_ascii=False, separators=(',', ':'))
        fh.write(';\n')
    print(f"цен по ОКП: {len(by_okp)}, по обозначению: {len(by_des)} -> {out}")


if __name__ == '__main__':
    main()
